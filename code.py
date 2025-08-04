from openpyxl import load_workbook
import os

workbook = load_workbook("Book.xlsx")
sheet = workbook.active

print(os.getcwd())

while True:
    val = '126225'+input("Last 4 digits of PRN: ")

    for i in range(2, sheet.max_row + 1):
        if sheet[f'B{i}'].value == int(val):
            print(f"{sheet[f'C{i}'].value} is Present")
            sheet[f'D{i}'].value = 'P'

    workbook.save('Book.xlsx')